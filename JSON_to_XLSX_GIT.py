import json
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError


def open_json(infile:str):
    with open(infile, 'r', encoding="UTF-8") as json_file:
        data = json.load(json_file)
    return data


def parse_json(data):
    keys = []
    for key in data[0].keys():
        keys.append(key)
    structure = []
    for element in data:
        source = element.values()
        row = []
        for item in source:
            row.append(item)
        structure.append(row)
    return keys, structure


def create_xlsx(keys, structure):
    new_wb = Workbook()
    ws = new_wb.active
    ws.append(keys)
    start_row = 2
    for i, row in enumerate(structure, start=start_row):
        for j, value in enumerate(row, start=1):  # Столбцы начинаются с 1
            try:
                ws.cell(row=i, column=j, value=value)
            except (ValueError, IllegalCharacterError, TypeError) as e:
                print(f"Json data cannot be converted to XLSX {value} in cell ({i}, {j}): {e}")
                ws.cell(row=i, column=j, value="Error")

    new_wb.save("friends.xlsx")  # name of your xlsx file


if __name__ == "__main__":
    data = open_json("XXX") # link to your json file
    keys, structure = parse_json(data)
    create_xlsx(keys, structure)



